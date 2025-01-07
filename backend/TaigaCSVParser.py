import os
import io

import pandas as pd
import openpyxl as opyxl

import requests

import tkinter as tk
from tkinter import filedialog

root = tk.Tk()
root.withdraw()

class TaigaParsingController:
    us_report_url = None
    task_report_url = None

    us_fp = None
    task_fp = None

    member_list = None
    sprint_list = None
    us_list = None
    raw_us_df = None
    raw_task_df = None
    formatted_master_df = None

    data_ready = False
    
    def __init__(self, us_url, task_url):
        self.set_us_report_url(us_url)
        self.set_task_report_url(task_url)

    ## API/File declaration
    ##=============================================================================

    def set_us_report_url(self, url):
        if url != "" and url is not None:
            self.us_report_url = url

    def get_us_report_url(self):
        return self.us_report_url

    def set_task_report_url(self, url):
        if url != "" and url is not None:
            self.task_report_url = url

    def get_task_report_url(self):
        return self.task_report_url

    def __file_is_valid(self, filename):
        if filename != "" and filename is not None:
            return True
        else:
            return False

    def set_us_fp(self, fp):
        if self.__file_is_valid(fp):
            self.us_fp = fp
            return True
        return False
    
    def get_us_fp(self):
        return self.us_fp

    def set_task_fp(self, fp):
        if self.__file_is_valid(fp):
            self.task_fp = fp
            return True
        return False
    
    def get_task_fp(self):
        return self.task_fp
    
    ## Retrieval of data
    ##=============================================================================

    def __us_data_from_api(self):
        url = self.us_report_url

        if url != "" and url is not None:
            res = requests.get(url)._content
            raw_data = pd.read_csv(io.StringIO(res.decode('utf-8')))
            return self.__format_us_data(raw_data)
        return False

    def __task_data_from_api(self):
        url = self.task_report_url

        if url != "" and url is not None:
            res = requests.get(url)._content
            raw_data = pd.read_csv(io.StringIO(res.decode('utf-8')))
            return self.__format_task_data(raw_data)
        return False

    def retrieve_data_by_api(self):
        task_parse_success = self.__task_data_from_api()
        us_parse_success = self.__us_data_from_api()

        if task_parse_success and us_parse_success:
            self.__format_and_centralize_data()
            self.data_ready = True
        else:
            print('Error retrieving and parsing data through api')

    def __us_data_from_file(self):
        us_fp = self.us_fp
        us_url_set = self.__file_is_valid(us_fp)

        if not us_url_set:
            us_url_set = self.set_us_fp()

        if us_url_set:
            raw_data = None

            if us_fp.split(".")[-1] == "csv":
                raw_data = pd.read_csv(us_fp)
            elif us_fp.split(".")[-1] == "xlsx":
                raw_data = pd.read_excel(us_fp)
            return self.__format_us_data(raw_data)
        return False
    
    def __parse_user_stories(self, df):
        user_stories = []

        for row in df:
            if pd.notnull(row):
                user_stories.append(row) if row not in user_stories else None

        self.us_list = user_stories
    
    def get_user_stories(self):
        return self.us_list

    def __task_data_from_file(self):
        task_fp = self.task_fp
        task_url_set = self.__file_is_valid(task_fp)

        if not task_url_set:
            task_url_set = self.set_task_fp()

        if task_url_set:
            raw_data = None

            if task_fp.split(".")[-1] == "csv":
                raw_data = pd.read_csv(task_fp)
            elif task_fp.split(".")[-1] == "xlsx":
                raw_data = pd.read_excel(task_fp)
            return self.__format_task_data(raw_data)
        return False

    def retrieve_data_by_file(self):
        task_parse_success = self.__task_data_from_file()
        us_parse_success = self.__us_data_from_file()

        if task_parse_success and us_parse_success:
            self.__format_and_centralize_data()
            self.data_ready = True
        else:
            print('Error retrieving and parsing data from files')

    def clear_data(self):
        self.member_list = None
        self.sprint_list = None
        self.raw_us_df = None
        self.raw_task_df = None
        self.formatted_master_df = None

    def __parse_members(self, dataframe):
        members = []

        for row in dataframe:
            if pd.notnull(row):
                members.append(row) if row not in members else None

        self.member_list = members

    def get_members(self):
        return self.member_list

    def __parse_sprints(self, df):
        sprints = []

        for row in df:
            if pd.notnull(row):
                sprints.append(row) if row not in sprints else None

        self.sprint_list = sprints

    def get_sprints(self):
        return self.sprint_list
    
    
    
    ## Data preparation
    ##=============================================================================
    
    def __format_us_data(self, raw_df):
        if raw_df is not None:
                self.raw_us_df = raw_df[['id', 'ref', 'sprint', 'total-points']]
                self.__parse_sprints(raw_df['sprint'])
                self.__parse_user_stories(raw_df['ref'])
                return True
        return False
    
    def __format_task_data(self, raw_df):
        if raw_df is not None:
            raw_df = raw_df[['id', 'ref', 'subject', 'user_story', 'sprint', 'assigned_to']]
            raw_df.sort_values(['sprint'], ascending=[True], inplace=True)
            self.raw_task_df = raw_df
            self.__parse_members(raw_df['assigned_to'])
            return True
        return False

    def __make_hyperlink(self, task_num):
        base_url = 'https://tree.taiga.io/project/dlsmallw-group-8-asu-capstone-natural-language-processing-for-decolonizing-harm-reduction-literature/task/{}'
        return '=HYPERLINK("%s", "Task-%s")' % (base_url.format(task_num), task_num)

    def __format_and_centralize_data(self):
        data_columns = ['sprint', 'user_story', 'points', 'task', 'assigned_to', 'coding', 'subject']

        us_df = self.raw_us_df
        task_df = self.raw_task_df
        print(task_df.columns)

        all_data = [0] * len(task_df)

        for index, row in task_df.iterrows():
            us_num = row['user_story']
            us_row = us_df.loc[us_df['ref'] == us_num]

            sprint = row['sprint']
            user_story = int(us_num) if pd.notnull(us_num) else 'Storyless'
            points = int(us_row['total-points'].values[0] if pd.notnull(us_num) else 0)
            task = int(row['ref'])
            assigned = row['assigned_to'] if pd.notnull(row['assigned_to']) else 'Unassigned'
            coding = ""
            subject = row['subject']
            
            data_row = [sprint, user_story, points, task, assigned, coding, subject]
            all_data[index] = data_row

        self.formatted_master_df = pd.DataFrame(all_data, columns=data_columns)

    def __format_df_for_excel(self, df):
        members = self.get_members()
        num_mems = len(members)

        data_columns = ['sprint', 'user_story', 'points', 'task', 'coding']
        data_columns.extend(members)

        data = [0] * len(df)
        for index, row in df.iterrows():
            us_num = row['user_story']
            task_num = row['task']
            assigned = row['assigned_to']

            sprint = row['sprint']
            user_story = f'US-{int(us_num)}' if us_num != 'Storyless' else 'Storyless'
            points = int(row['points'])
            task = self.__make_hyperlink(task_num)
            coding = row['coding']

            mem_data = [None] * num_mems
            i = 0
            for mem in members:
                mem_data[i] = "100%" if assigned == mem else None
                i += 1

            row_data = [sprint, user_story, points, task, coding]
            row_data.extend(mem_data)
            data[index] = row_data
            
        result_df = pd.DataFrame(data, columns=data_columns)
        return result_df
    

    
    ## File Writing
    ##=============================================================================

    def __create_new_wb(self, filename):
        if (os.path.exists(filename)):
            os.remove(filename)
        
        wb = opyxl.Workbook()
        wb.create_sheet("All_Data")

        for sprint in self.sprint_list:
            wb.create_sheet(sprint)

        for sheet in wb.sheetnames:
            if sheet not in self.sprint_list and sheet != "All_Data":
                del wb[sheet]

        wb.save(filename)

    def __parsed_data_to_spreadsheet(self, df, writer, sheet):
        df.to_excel(writer, sheet_name=sheet)

    def write_data(self, filename):
        self.__create_new_wb(filename, self.sprint_list)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            master_df = self.get_master_df()
            self.__parsed_data_to_spreadsheet(
                self.__format_df_for_excel(master_df), writer, "All_Data")
            
            for sprint in self.sprint_list:
                sprint_df = master_df.loc[master_df['sprint'] == sprint]
                self.__parsed_data_to_spreadsheet(
                    self.__format_df_for_excel(sprint_df.sort_values(['sprint', 'user_story', 'task'], ascending=[True, True, True])), 
                    writer, sprint)
                
    def get_master_df(self):
        return self.formatted_master_df.sort_values(['sprint', 'user_story', 'task'], ascending=[True, True, True])