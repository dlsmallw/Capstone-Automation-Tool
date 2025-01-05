
import os
import pandas as pd
import requests
import datetime
import openpyxl as opyxl
import pytz
import re

class GitHubParsingController:
    config_fp = os.path.join(os.getcwd(), 'config.txt')
    config_parser = None
    gh_base_url = "https://api.github.com"

    gh_username = None
    gh_token = None
    gh_repo_owner = None
    gh_repo_name = None

    contributor_list = None
    branch_list = None
    raw_commit_df = None
    commits_by_committer_df = None
    latest_commit_date = None

    api_ref_validated = False
    auth_verified = False
    data_ready = False

    def __init__(self, username, token, owner, repo):
        self.set_gh_auth(username, token)
        self.set_repo_details(owner, repo)
        self.__load_from_csv()

    ## Auth Management
    ##=============================================================================
    
    def validate_username(self, username):
        return username != "" and username is not None
    
    def validate_token(self, token):
        return token != "" and token is not None

    def validate_auth(self, username, token):
        if self.validate_username(username) and self.validate_token(token):
            res = self.__make_gh_api_call(f'{self.gh_base_url}/user')
            if res.status_code >= 200 and res.status_code < 300:
                self.auth_verified = True
                return True
        return False
    
    def auth_validated(self):
        return self.auth_verified

    def set_gh_username(self, username):
        if self.validate_username(username):
            self.gh_username = username
            return True
        return False
    
    def get_username(self):
        return self.gh_username

    def set_gh_token(self, token):
        if self.validate_token(token):
            self.gh_token = token
            return True
        return False
    
    def get_token(self):
        return self.gh_token

    def set_gh_auth(self, username=None, token=None):
        if username is not None:
            username_set = self.set_gh_username(username)
            if not username_set:
                print('Invalid username provided')
        else:
            if self.gh_username is not None:
                username_set = True

        if token is not None:
            token_set = self.set_gh_token(token)
            if not token_set:
                print('Invalid token provided')
        else:
            if self.gh_token is not None:
                token_set = True

        if username_set and token_set:
            return self.validate_auth(username, token)
        else:
            return False
        
    ## Target Repository Management
    ##=============================================================================
        
    def validate_repo_exists(self):
        owner = self.gh_repo_owner
        repo = self.gh_repo_name

        url = f'{self.gh_base_url}/repos/{owner}/{repo}'
        header = self.__get_auth_header()

        res = requests.get(url, headers=header)

        if res.status_code >= 200 and res.status_code < 300:
            self.api_ref_validated = True
            return True
        return False
    
    def repo_validated(self):
        return self.api_ref_validated

    def set_repo_owner_username(self, owner):
        if self.validate_username(owner):
            self.gh_repo_owner = owner
            return True
        return False
    
    def get_repo_owner(self):
        return self.gh_repo_owner
    
    def set_repo_name(self, repo):
        if repo != "" and repo is not None:
            self.gh_repo_name = repo
            return True
        return False
    
    def get_repo_name(self):
        return self.gh_repo_name
    
    def set_repo_details(self, owner, repo):
        repo_details_success = True

        if not self.set_repo_owner_username(owner):
            print('Invalid repo owner username')
            repo_details_success = False

        if not self.set_repo_name(repo):
            print('Invalid repo name')
            repo_details_success = False

        if repo_details_success:
            return self.validate_repo_exists()
        else:
            return repo_details_success
        
    ## API Call Management
    ##=============================================================================
        
    def __get_auth_header(self):
        return {
            'Authorization': f'token {self.gh_token}' 
        }
    
    def __make_gh_api_call(self, url):
        header = self.__get_auth_header()
        return requests.get(url, headers=header)
    
    def __get_commit_author(self, commit):
        author_name = commit["author"]["login"]
        author_email = commit["commit"]["author"]["email"]

        if author_name is not None and author_name != 'unknown':
            if author_name in self.contributor_list:
                return author_name
        else:
            if author_email is not None and author_email != 'unknown':
                suspected_name = author_email[0:author_email.index('@')]

                if suspected_name in self.contributor_list:
                    return suspected_name
        return 'unknown'
    
    def __get_paginated_branch_data(self, url):
        pagesRemaining = True
        branches = dict()
        next_url = url

        while pagesRemaining:
            res = self.__make_gh_api_call(next_url)
            links = res.links
            data = res.json()

            for entry in data:
                name = entry['name']
                last_commit_sha = entry['commit']['sha']
                branches[name] = last_commit_sha
            
            try:
                next_url = links.get('next').get('url')
            except:
                pagesRemaining = False

        return branches
    
    def __get_paginated_commit_data(self, url):
        pagesRemaining = True
        commits = []
        next_url = url

        while pagesRemaining:
            res = self.__make_gh_api_call(next_url)
            links = res.links
            data = res.json()
            pattern = r'task[^a-zA-Z\d\s]?\d+'

            for commit_entry in data:
                commit_obj = commit_entry["commit"]

                # Author Details used to assign the commit
                committer = self.__get_commit_author(commit_entry)
                # Commit date and title
                commit_msg = commit_obj['message']

                if not 'merge' in commit_msg.lower():
                    match = re.search(pattern, commit_msg, re.IGNORECASE)

                    if match:
                        task_num = int(re.search(r'\d+', match.group()).group())
                    else:
                        task_num = -1
                
                    # Takes the commit timezone (UTC) and converts to AZ timezone
                    utc_dt = datetime.datetime.strptime(commit_obj['committer']['date'], '%Y-%m-%dT%H:%M:%SZ')
                    az_dt = utc_dt.astimezone(pytz.timezone('US/Arizona'))
                    
                    # Used for identifying and filtering commits
                    id = commit_entry['sha']
                    url = commit_entry['html_url']

                    commits.append({
                        "id": id,
                        "task_num": task_num,
                        "url": url,
                        "message": commit_msg,
                        "utc_datetime": f'{utc_dt}',
                        "az_date": f'{az_dt.strftime('%m/%d/%Y')}',
                        "committer": committer
                    })

            try:
                next_url = links.get('next').get('url')
            except:
                pagesRemaining = False
        
        raw_commits_data = pd.json_normalize(commits)
        return raw_commits_data
    
    def __parse_repo_branches(self):
        owner = self.gh_repo_owner
        repo = self.gh_repo_name
        url = f'{self.gh_base_url}/repos/{owner}/{repo}/branches?per_page=100'
        self.branch_list = self.__get_paginated_branch_data(url)

    def get_branches(self):
        return list(self.branch_list.keys())
    
    def __parse_repo_contributors(self):
        owner = self.gh_repo_owner
        repo = self.gh_repo_name
        url = f'{self.gh_base_url}/repos/{owner}/{repo}/contributors'

        res = self.__make_gh_api_call(url).json()

        contributors = []

        for entry in res:
            contributors.append(entry['login'])

        contributors.append('unknown')
        self.__set_contributors(contributors)

    def __set_contributors(self, contributors):
        self.contributor_list = contributors

    def get_contributors(self):
        return self.contributor_list
    
    def __parse_all_commits(self):
        owner = self.gh_repo_owner
        repo = self.gh_repo_name
        since = self.latest_commit_date

        all_data = self.get_all_commit_data()

        for branch in self.get_branches():
            branch_sha = self.branch_list[branch]

            if since:
                url = f'{self.gh_base_url}/repos/{owner}/{repo}/commits?since={since}&per_page=100&sha={branch_sha}'
            else:
                url = f'{self.gh_base_url}/repos/{owner}/{repo}/commits?per_page=100&sha={branch_sha}'

            print(f'{branch}: {branch_sha}')

            branch_commits = self.__get_paginated_commit_data(url)

            if all_data is None:
                all_data = branch_commits
            else:
                all_data = pd.concat([all_data, branch_commits]).drop_duplicates(subset=['id'], keep='first').reset_index(drop=True)

        self.__set_commit_data(all_data)
    
    def __set_commit_data(self, all_data):
        all_data['utc_datetime'] = pd.to_datetime(all_data['utc_datetime'])
        all_data.sort_values(by='utc_datetime', inplace=True)
        latest = all_data['utc_datetime'].max().date()
        self.raw_commit_df = all_data
        self.latest_commit_date = f'{latest.isoformat()}T00:00:00Z'
        print(self.latest_commit_date)

    def get_all_commit_data(self):
        return self.raw_commit_df
    
    def __parse_commits_by_committer(self):
        commit_data = self.get_all_commit_data()
        contributors = self.get_contributors()
        commits_by_committer = dict()

        for contributor in contributors:
            commiter_df = commit_data.loc[commit_data['committer'] == contributor]
            commits_by_committer[contributor] = commiter_df

        self.commits_by_committer_df = commits_by_committer

    def get_commits_by_committer_data(self):
        return self.commits_by_committer_df
    
    ## Handling of data
    ##=============================================================================
    
    def retrieve_and_parse_commit_data(self):
        self.__parse_repo_contributors()
        if len(self.get_contributors()) <= 0:
            return
        
        self.__parse_repo_branches()
        if len(self.get_branches()) <= 0:
            return

        self.__parse_all_commits()
        if len(self.get_all_commit_data()) <= 0:
            return
        self.__store_raw_data()
        
        self.__parse_commits_by_committer()
        if len(self.get_commits_by_committer_data()) <= 0:
            return
        
        self.data_ready = True

    def clear_data(self):
        self.contributor_list = None
        self.raw_commit_df = None
        self.commits_by_committer_df = None
        self.data_ready = False

    ## File Writing
    ##=============================================================================
            
    def __create_new_wb(self, filename, sheets):
        if os.path.exists(filename):
            os.remove(filename)
        
        wb = opyxl.Workbook()
        wb.create_sheet("All_Data")

        for contributor in sheets:
            wb.create_sheet(contributor)

        for sheet in wb.sheetnames:
            if sheet not in self.contributor_list and sheet != "All_Data":
                del wb[sheet]

        wb.save(filename)
        wb.close()

    def __parsed_data_to_spreadsheet(self, df, writer, sheet):
        df.to_excel(writer, sheet_name=sheet, index=False)

    def write_to_csv(self, filename='./data_out/commit_data.csv'):
        if 'data_out' in filename:
            dir_name = 'data_out'
        else:
            dir_name = 'raw_data'

        if not os.path.exists(f'./{dir_name}'):
            os.makedirs(f'./{dir_name}')
        self.__write_data(filename, False)

    def __load_from_csv(self):
        filename = './raw_data/raw_commit_data.csv'
        if os.path.exists(filename):
            print(' > LOADING COMMIT DATA FROM CSV FILE')
            all_data = pd.read_csv(filename)
            contributors = sorted(all_data['committer'].unique())
            self.__set_commit_data(all_data)
            self.__set_contributors(contributors)
            self.__parse_commits_by_committer()
        else:
            print(' > NO COMMIT DATA CSV FILE')

    def __store_raw_data(self):
        self.write_to_csv('./raw_data/raw_commit_data.csv')

    def write_to_excel(self, filename='./data_out/commit_data.xlsx'):
        if 'data_out' in filename:
            dir_name = 'data_out'
        else:
            dir_name = 'raw_data'

        if not os.path.exists(f'./{dir_name}'):
            os.makedirs(f'./{dir_name}')

        self.__write_data(filename)

    def __write_data(self, filename, excel=True):
        if excel:
            contributors = self.get_contributors()
            all_data = self.get_all_commit_data()[['id', 'task_num', 'committer', 'message', 'az_date', 'url']]
            commits_by_contributor = self.get_commits_by_committer_data()

            self.__create_new_wb(filename, contributors)
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                self.__parsed_data_to_spreadsheet(all_data, writer, 'All_Data')

                for contributor in contributors:
                    contributor_df = commits_by_contributor[contributor][['id', 'task_num', 'message', 'az_date', 'url']]
                    self.__parsed_data_to_spreadsheet(contributor_df, writer, contributor)
        else:
            all_data = self.get_all_commit_data()[['id', 'task_num', 'committer', 'message', 'utc_datetime', 'az_date', 'url']]
            all_data.to_csv(filename, index=False)