import configparser
import os
import openpyxl as opyxl
import pandas as pd
from typing import Type
from backend.TaigaCSVParser import TaigaParsingController
from backend.GitHubCommitParser import GitHubParsingController

class DataController:
    config_fp = os.path.join(os.getcwd(), 'config.txt')
    config_parser = None

    tp = None
    ghp = None

    gh_auth_verified = False
    gh_repo_verified = False

    def __init__(self):
        self.__load_config()
        self.__load_raw_data()

    ## Config Management
    ##=============================================================================

    def __load_gh_config(self):
        config = self.config_parser

        if config.has_section('github-config'):
            gh_username = config.get('github-config', 'gh_username')
            gh_token = config.get('github-config', 'gh_token')
            repo_owner = config.get('github-config', 'gh_repo_owner')
            repo_name = config.get('github-config', 'gh_repo_name')

            self.ghp = GitHubParsingController(gh_username, gh_token, repo_owner, repo_name)
        else:
            self.__build_gh_section()

    def __load_taiga_config(self):
        config = self.config_parser

        if config.has_section('taiga-config'):
            us_report_url = config.get('taiga-config', 'us_report_api_url')
            task_report_url = config.get('taiga-config', 'task_report_api_url')

            self.tp = TaigaParsingController(us_report_url, task_report_url)
        else:
            self.__build_taiga_section()

    def __load_config(self):
        config = configparser.RawConfigParser()

        if not os.path.exists(self.config_fp):
            open(self.config_fp, 'w').close()
            self.__build_config_file()            
        config.read(self.config_fp)
        self.config_parser = config  # for future use
        self.__load_gh_config()
        self.__load_taiga_config()

    def __build_gh_section(self):
        config = self.config_parser
        config.add_section('github-config')
        self.__update_option_in_config('github-config', 'gh_username', None)
        self.__update_option_in_config('github-config', 'gh_token', None)
        self.__update_option_in_config('github-config', 'gh_repo_owner', None)
        self.__update_option_in_config('github-config', 'gh_repo_name', None)

    def __build_taiga_section(self):
        config = self.config_parser
        config.add_section('taiga-config')
        self.__update_option_in_config('taiga-config', 'us_report_api_url', None)
        self.__update_option_in_config('taiga-config', 'task_report_api_url', None)

    def __build_config_file(self):
        self.__build_taiga_section()
        self.__build_gh_section()

    def __update_gh_config_opt(self, option, value):
        self.__update_option_in_config('github-config', option, value)

    def __update_taiga_config_opt(self, option, value):
        self.__update_option_in_config('taiga-config', option, value)

    def __update_option_in_config(self, section, option, value):
        config = self.config_parser
        config.set(section, option, value)
        with open(self.config_fp, 'w') as configfile:
            config.write(configfile)
            configfile.close()

    def clear_config(self):
        self.__update_taiga_config_opt('us_report_api_url', None)
        self.__update_taiga_config_opt('task_report_api_url', None)
        self.__update_gh_config_opt('gh_username', None)
        self.__update_gh_config_opt('gh_token', None)
        self.__update_gh_config_opt('gh_repo_owner', None)
        self.__update_gh_config_opt('gh_repo_name', None)

    ## GH/Taiga Controller Management
    ##=============================================================================

    def set_gh_auth(self, username, token):
        if self.ghp.set_gh_auth(username, token):
            self.gh_auth_verified = self.ghp.auth_validated()
            self.__update_gh_config_opt('gh_username', username)
            self.__update_gh_config_opt('gh_token', token)
    
    def set_gh_username(self, username):
        success = self.ghp.set_gh_username(username)
        if success:
            self.__update_gh_config_opt('gh_username', username)
            token = self.ghp.get_token()
            if self.ghp.validate_token(token):
                self.gh_auth_verified = self.ghp.set_gh_auth(username, token)
        return success

    def get_gh_username(self):
        return self.ghp.get_username()
    
    def set_gh_token(self, token):
        success = self.ghp.set_gh_token(token)
        if success:
            self.__update_gh_config_opt('gh_token', token)
            username = self.ghp.get_username()
            if self.ghp.validate_username(username):
                self.gh_auth_verified = self.ghp.set_gh_auth(username, token)
        return success
    
    def get_gh_token(self):
        return self.ghp.get_token()
    
    def set_gh_repo_details(self, owner, repo):
        if self.ghp.set_repo_details(owner, repo):
            self.gh_repo_verified = self.ghp.repo_validated()
            self.__update_gh_config_opt('gh_repo_owner', owner)
            self.__update_gh_config_opt('gh_repo_name', repo)

    def set_gh_owner(self, owner):
        success = self.ghp.set_repo_owner_username(owner)
        if success:
            self.__update_gh_config_opt('gh_repo_owner', owner)
            self.gh_repo_verified = self.ghp.validate_repo_exists()

    def get_repo_owner(self):
        return self.ghp.get_repo_owner()
    
    def set_gh_repo(self, repo):
        success = self.ghp.set_repo_name(repo)
        if success:
            self.__update_gh_config_opt('gh_repo_name', repo)
            self.gh_repo_verified = self.ghp.validate_repo_exists()


    def get_repo_name(self):
        return self.ghp.get_repo_name()
    
    def set_taiga_us_api_url(self, url):
        self.tp.set_us_report_url(url)
        self.__update_taiga_config_opt('us_report_api_url', url)

    def get_taiga_us_api_url(self):
        return self.tp.get_us_report_url()

    def set_taiga_task_url(self, url):
        self.tp.set_task_report_url(url)
        self.__update_taiga_config_opt('task_report_api_url', url)

    def get_taiga_task_api_url(self):
        return self.tp.get_task_report_url()

    def set_us_fp(self, fp):
        self.tp.set_us_fp(fp)

    def get_us_fp(self):
        return self.tp.get_us_fp()

    def set_task_fp(self, fp):
        self.tp.set_task_fp(fp)

    def get_task_fp(self):
        return self.tp.get_task_fp()

    ## API calling
    ##=============================================================================

    def github_retrieve_and_parse(self):
        self.ghp.retrieve_and_parse_commit_data()

    def taiga_retrieve_from_api(self):
        self.tp.retrieve_data_by_api()

    def taiga_retrieve_from_files(self):
        self.tp.retrieve_data_by_file()
    
    ## Data retrieval/setting
    ##=============================================================================

    def get_members(self) -> list:
        return self.tp.get_members()
    
    def get_sprints(self) -> list:
        return self.tp.get_sprints()
    
    def get_user_stories(self) -> list:
        return self.tp.get_user_stories()

    def get_taiga_master_df(self):
        return self.tp.get_master_df()
    
    def get_gh_master_df(self) -> pd.DataFrame:
        return self.ghp.get_all_commit_data()
    
    def set_taiga_master_df(self, df):
        self.tp.set_master_df(df)

    def clear_taiga_data(self):
        self.tp.clear_data()
        self.remove_file('./raw_data/raw_taiga_master_data.csv')
        self.remove_file('./raw_data/raw_taiga_us_data.csv')
        self.remove_file('./raw_data/raw_taiga_task_data.csv')
    
    def taiga_data_ready(self):
        return self.tp.data_is_ready()
    
    def github_data_ready(self):
        return self.ghp.data_is_ready()
    
    ## Raw Data Storage/Loading
    ##=============================================================================

    def __create_new_wb(self, filename, sheets):
        if os.path.exists(filename):
            os.remove(filename)
        
        wb = opyxl.Workbook()
        wb.create_sheet("All_Data")

        for contributor in sheets:
            wb.create_sheet(contributor)

        for sheet in wb.sheetnames:
            if sheet not in self.contributor_list and sheet != "All_Data":
                del wb[sheet]

        wb.save(filename)
        wb.close()

    def __parsed_data_to_spreadsheet(self, df, writer, sheet):
        df.to_excel(writer, sheet_name=sheet, index=False)

    def write_to_csv(self, filepath: Type[str], df: Type[pd.DataFrame]):
        subdirectories = filepath.replace('.', '').split('/')
        curr_dir_level = '.'
        for i in range(len(subdirectories) - 1):
            dir = subdirectories[i]
            curr_dir_level += f'/{dir}'

            if not os.path.exists(curr_dir_level):
                os.makedirs(curr_dir_level)

        df.to_csv(filepath, index=False)

    def __load_from_csv(self, filepath) -> pd.DataFrame | None:
        df = None
        if os.path.exists(filepath):
            print(f' > Loading data from {filepath}')
            df = pd.read_csv(filepath)
            df.replace(['', 'None', 'nan', 'NaN'], [None, None, None, None], inplace=True)
        else:
            print(f' > File {filepath} does not exist')
        return df
    
    def __load_raw_data(self):
        taiga_raw_master_df = self.__load_from_csv('./raw_data/raw_taiga_master_data.csv')
        taiga_raw_us_df = self.__load_from_csv('./raw_data/raw_taiga_us_data.csv')
        taiga_raw_task_df = self.__load_from_csv('./raw_data/raw_taiga_task_data.csv')
        self.tp.load_raw_data(taiga_raw_master_df, taiga_raw_us_df, taiga_raw_task_df)

        github_df = self.__load_from_csv('./raw_data/raw_github_data.csv')
        self.ghp.load_raw_data(github_df)

    def __store_all_raw_data(self, filename, df):
        self.write_to_csv(filename, df)

    def store_raw_taiga_data(self):
        raw_master_df = self.tp.get_master_df()
        raw_us_df = self.tp.get_raw_us_data()
        raw_task_df = self.tp.get_raw_task_data()

        if raw_master_df is not None:
            self.write_to_csv('./raw_data/raw_taiga_master_data.csv', raw_master_df)

        if raw_us_df is not None:
            self.write_to_csv('./raw_data/raw_taiga_us_data.csv', raw_us_df)

        if raw_task_df is not None:
            self.write_to_csv('./raw_data/raw_taiga_task_data.csv', raw_task_df)

    def __store_raw_github_data(self, df):
        self.write_to_csv('./raw_data/raw_github_data.csv', df)

    def write_to_excel(self, filename='./data_out/commit_data.xlsx'):
        if 'data_out' in filename:
            dir_name = 'data_out'
        else:
            dir_name = 'raw_data'

        if not os.path.exists(f'./{dir_name}'):
            os.makedirs(f'./{dir_name}')

        self.__write_data(filename)

    def __write_data(self, filename, excel=True):
        if excel:
            contributors = self.get_contributors()
            all_data = self.get_all_commit_data()[['id', 'task_num', 'committer', 'message', 'az_date', 'url']]
            commits_by_contributor = self.get_commits_by_committer_data()

            self.__create_new_wb(filename, contributors)
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                self.__parsed_data_to_spreadsheet(all_data, writer, 'All_Data')

                for contributor in contributors:
                    contributor_df = commits_by_contributor[contributor][['id', 'task_num', 'message', 'az_date', 'url']]
                    self.__parsed_data_to_spreadsheet(contributor_df, writer, contributor)
        else:
            all_data = self.get_all_commit_data()[['id', 'task_num', 'committer', 'message', 'utc_datetime', 'az_date', 'url']]
            all_data.to_csv(filename, index=False)

    def remove_file(self, filepath):
        if os.path.exists(filepath):
            os.remove(filepath)

