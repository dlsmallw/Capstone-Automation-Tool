import openpyxl as opyxl
from openpyxl import styles
import os
import datetime
import pandas as pd

def format_spreadsheet(filepath):
    if not os.path.exists(filepath):
        return
    
    try:
        wb = opyxl.load_workbook(filepath)
        for sheet in wb.worksheets:
            col_dims = dict()
            colidx_range = [i for i in range(sheet.min_column, sheet.max_column + 1)]
            for colidx in colidx_range:
                column = sheet.cell(1, colidx).column_letter
                link_col = False
                date_col = False
                int_col = False

                header = sheet.cell(1, colidx).value
                if 'link' in header.lower():
                    link_col = True
                elif 'date' in header.lower():
                    date_col = True
                elif header in ['Task #', 'User Story', 'Points']:
                    int_col = True

                col_dims[column] = len(str(header))
                for i in range(2, sheet.max_row + 1):
                    cell = sheet.cell(i, colidx)
                    
                    if link_col:
                        if isinstance(cell.value, str) and cell.value.startswith('=HYPERLINK('):
                            try:
                                # Extract URL from the formula
                                url_start = cell.value.find('"') + 1
                                url_end = cell.value.find('"', url_start)
                                url = cell.value[url_start:url_end] if url_start > 0 and url_end > url_start else ""

                                if url is not None and url != '':
                                    # Extract Friendly Text (if available)
                                    text_start = cell.value.find('"', url_end + 1) + 1
                                    text_end = cell.value.find('"', text_start)
                                    friendly_text = cell.value[text_start:text_end] if text_start > 0 and text_end > text_start else url

                                    # Set the hyperlink in the cell
                                    cell.value = friendly_text  # Display text
                                    cell.hyperlink = url  # Set hyperlink
                                    cell.font = styles.Font(underline='single', color='0000FF') # Apply Excel hyperlink style 

                            except Exception as e:
                                print(f"Error processing cell {cell.coordinate}: {e}")
                    elif date_col:
                        date = pd.to_datetime(cell.value)
                        cell.value = date
                        cell.number_format = 'm/d/yyyy'
                    elif int_col:
                        try:
                            cell.value = int(cell.value)
                            cell.number_format = styles.numbers.FORMAT_NUMBER
                        except: 
                            pass

                    col_dims[column] = max(col_dims.get(column), len(str(cell.value)))
            for col, val in col_dims.items():
                sheet.column_dimensions[col].width = val
        wb.save(filepath)

    except Exception as e:
        print(f'Failed to format spreadsheet - {e}')

fp = r"C:\Users\Daniel\Desktop\Work_Summary_Report.xlsx"
format_spreadsheet(fp)