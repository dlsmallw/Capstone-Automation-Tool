import configparser
import os

import openpyxl as opyxl
import pandas as pd
import numpy as np
from typing import Type
from openpyxl import styles
from models.Taiga import TaigaDataServicer
from models.GitServicerInterface import GitServicer
from models.database.RecordDatabase import RecDB
import requests
import http.client as hc

TAIGA = 'Taiga'
GITHUB = 'GitHub'
GITLAB = 'GitLab'

class DataController:
    config_fp = os.path.join(os.getcwd(), 'config.txt')
    config_parser = None

    gh_auth_verified = False
    gh_repo_verified = False

    def __init__(self, db: RecDB):
        ## Class object defaults
        self.db : RecDB = None

        ## Data Servicers
        self.ts : TaigaDataServicer = None
        self.gs : GitServicer = None

        ## Taiga-related variables
        #### Projects
        self.taiga_projects_df : pd.DataFrame = None
        self.sel_pid = None
        self.sel_project_name: str = None
        self.sel_project_owner: str = None
        self.sel_project_slug: str = None
        self.project_selected = False
        #### Dataframes
        self.sprints_df : pd.DataFrame = None
        self.members_df : pd.DataFrame = None
        self.us_df : pd.DataFrame = None
        self.tasks_df : pd.DataFrame = None
        #### CSV URLs
        self.us_report_url = None
        self.task_report_url = None
        #### Status Vars
        self.taiga_data_available = False

        ## Git-related variables
        self.git_accts : dict = dict()
        self.repos : pd.DataFrame = None
        self.contributors : list[str] = None
        self.commits_df : pd.DataFrame = None
        
        self.commit_data_available = False

        ## Instance initializations and data loading
        self.db = db
        self.ts = self._init_taiga_servicer()
        self.gs = self._init_git_servicer()

    ## Initialization Functions
    ##=============================================================================

    def _init_taiga_servicer(self) -> TaigaDataServicer | None:
        def load_taiga_projects():
            projects = self.db.table_to_df('taiga_projects')
            if projects is not None and len(projects) > 0:
                self.taiga_projects_df = projects
                row = projects.loc[projects['is_selected'] == 1].values[0]

                if len(row) > 0:
                    self._set_linked_taiga_project(row[0], row[1], row[2], row[3])

        def load_saved_taiga_data():
            self._update_sprints_df(self.db.table_to_df('sprints'))
            self._update_members_df(self.db.table_to_df('members'))
            self.update_us_df(self.db.table_to_df('userstories'))
            self.update_tasks_df(self.db.table_to_df('tasks'))

        load_taiga_projects()
        load_saved_taiga_data()
        return TaigaDataServicer(self.load_taiga_credentials())
    
    def _init_git_servicer(self) -> GitServicer:
        def load_accts():
            accts = self.db.table_to_df('sites')
            repo_accts = accts[accts['nickname'] != 'Taiga']
            for index, row in repo_accts.iterrows():
                user, token = self.db.decrypt([row['username'], row['site_token']])
                repo_accts.loc[index, 'username'] = user
                repo_accts.loc[index, 'site_token'] = token
            
            if repo_accts is not None and len(repo_accts) > 0:
                self.git_accts = dict()

                for index, row in repo_accts.iterrows():
                    site = row['site_name']
                    nickname = row['nickname']
                    user = row['username']
                    token = row['site_token']

                    res, msg = self._validate_token(site, token)

                    if res == 'Success':
                        details = "Ready to make API calls"
                        gs.init_git_servicer(site, nickname, token)
                    else:
                        details = msg

                    self.git_accts[nickname] = {
                        'site': site,
                        'user': user,
                        'token': token,
                        'details': details
                    }
        
        def load_repos():
            repos = self.db.table_to_df('repos')
            if repos is not None and len(repos) > 0:
                self.repos = repos
        
        def load_commit_data():
            self.update_commit_df(self.db.table_to_df('commits'))

        gs = GitServicer()
        load_repos()
        load_accts()
        load_commit_data()
        return gs
    

    ## Util Functions
    ##=============================================================================

    def _validate_token(self, site, token):
        def gh_request(token):
            base_url = "https://api.github.com/user"
            header = {
                'Accept': 'application/vnd.github+json',
                'Authorization': f'Bearer {token}',
                'X-GitHub-Api-Version': '2022-11-28'
            }

            return base_url, header
        
        def gl_request(token):
            base_url = 'https://gitlab.com/api/v4/user'
            header = {
                'PRIVATE-TOKEN': f'{token}'
            }

            return base_url, header

        if site == GITHUB:
            url, headers = gh_request(token)
            login_key = 'login'
        else:
            url, headers = gl_request(token)
            login_key = 'username'

        try:
            res = requests.get(url=url, headers=headers)
            sc = res.status_code

            if sc >= 200 and sc <= 200:
                return 'Success', res.json()[login_key]
            else:
                return 'Error', f'{sc} - [{hc.responses[sc]}]'
        except Exception as e:
            return 'Error', f'Exception - {e}'

    def _get_site_credentials_from_db(self, site_name):
        return self.db.decrypt(tuple(self.db.select('sites', ['username', 'user_pwd', 'site_token'], {'site_name': site_name})[0]))
    
    def _update_user_credentials(self, site_name, username=None, pwd=None, token=None):
        return self.db.update('sites', dict(zip(['username', 'user_pwd', 'site_token'], self.db.encrypt([username, pwd, token]))), {'site_name': site_name})
    
    def check_if_nickname_exists(self, nickname):
        result = self.db.select('sites', conditions={'nickname': nickname})
        return len(result) > 0
    
    def _check_if_token_exists(self, token):
        for acct in self.git_accts.keys():
            if token == self.git_accts[acct]['token']:
                return True
        return False
    
    def update_git_acct(self, site, nickname, token):
        res, msg = self._validate_token(site, token)
        if res == 'Error':
            return res, msg
        
        if self._check_if_token_exists(token):
            return 'Error', 'Account associated with this token already exists'
        
        username = msg
        enc_uname, enc_token = self.db.encrypt([msg, token])
        data = {
            'site_name': site,
            'username': enc_uname,
            'nickname': nickname,
            'site_token': enc_token
        }

        cond = {'nickname': nickname}
        self.db.insert('sites', data, cond)

        if nickname in self.git_accts.keys():
            self.git_accts[nickname]['user'] = username
            self.git_accts[nickname]['token'] = token
            self.git_accts[nickname]['details'] = "Ready to make API calls"
        else:
            self.git_accts[nickname] = {
                'site': site,
                'user': username,
                'token': token,
                'details': "Ready to make API calls"
            }

        return res, msg
    
    def add_git_acct(self, site, nickname, token):
        res, msg = self.update_git_acct(site, nickname, token)
        if res == 'Success':
            self.gs.init_git_servicer(site, nickname, token)
        return res, msg
    
    def remove_git_acct(self, nickname):
        print(f'Deleting acct {nickname}')
        self.gs.remove_servicer(nickname)
        self.git_accts.pop(nickname)
        self.repos = self.repos[self.repos['site_nickname'] != nickname]
        self.db.delete('sites', conditions={'nickname': nickname})
        self.db.delete('repos', conditions={'site_nickname': nickname})
    
    def get_git_accts(self):
        accts = []
        if self.git_accts is not None and len(self.git_accts) > 0:
            for nname in self.git_accts:
                site = self.git_accts[nname]['site']
                user = self.git_accts[nname]['user']
                details = self.git_accts[nname]['details']
                accts.append([site, nname, user, details])
        return accts
    
    def _get_acct_repos(self, nickname):
        self.pull_repos(nickname)
        return self.get_avail_repos()
    
    def link_repo(self, repo):
        repos = self.repos.copy(deep=True)
        repos.loc[repos['repo_name'] == repo, 'is_linked'] = 1
        self._update_repos(repos, cols=['is_linked'])

    def unlink_repo(self, repo):
        repos = self.repos.copy(deep=True)
        repos.loc[repos['repo_name'] == repo, 'is_linked'] = 0
        self._update_repos(repos, cols=['is_linked'])

    def repos_available(self) -> bool:
        return self.repos is not None and len(self.repos) > 0
    
    def repos_linked(self) -> bool:
        return len(self.get_linked_repos()) > 0
    
    def api_call_ready(self) -> bool:
        if self.repos_linked:
            for acct in self.git_accts.keys():
                if self.git_accts[acct]['details'] == "Ready to make API calls":
                    return True
        return False
        
    
    #### Taiga
    ####===========================================================================

    def get_us_df(self) -> pd.DataFrame:
        return self.us_df
    
    def get_task_df(self) -> pd.DataFrame:
        return self.tasks_df
    
    def get_sprints_df(self) -> pd.DataFrame:
        return self.sprints_df

    def get_num_projects(self):
        if self.taiga_projects_df is not None and not self.taiga_projects_df.empty:
            return len(self.taiga_projects_df)
        else:
            return 0
        
    def get_available_projects(self):
        projects = []
        
        for _, row in self.taiga_projects_df.iterrows():
            projects.append(row['project_name'])
        return projects
    
    def get_linked_taiga_project(self):
        return self.sel_pid, self.sel_project_name, self.sel_project_owner

    def _set_linked_taiga_project(self, pid, name, owner, slug):
        if pid is not None and name is not None and owner is not None:
            self.sel_pid = pid
            self.sel_project_name = name
            self.sel_project_owner = owner
            self.sel_project_slug = slug
            self.project_selected = True

    def _get_taiga_base_url(self):
        base_url = None
        if self.project_selected:
            base_url = f'https://tree.taiga.io/project/{self.sel_project_slug}/task/'
        return base_url

    # Function to authenticate with Taiga API
    def authenticate_with_taiga(self, username, password):
        if not username or not password:
            return "Error", "Please enter both username and password."
            
        url = "https://api.taiga.io/api/v1/auth"  # Change this to your Taiga API URL if self-hosted
        data = {
            "type": "normal",
            "username": username,
            "password": password
        }

        try:
            response = requests.post(url, json=data)
            if response.status_code == 200:
                auth_token = response.json().get("auth_token")
                self.ts.update_user_credentials(username, password, auth_token)
                self._update_user_credentials('Taiga', username, password, auth_token)

                return "Success", f"Login successful! Token: {auth_token}"
            else:
                return "Error", f"Login failed: {response.json().get('non_field_errors', 'Unknown error')}"
        except Exception as e:
            return "Error", f"An error occurred: {e}"

    def select_taiga_project(self, project):
        sel_row = self.taiga_projects_df.loc[self.taiga_projects_df['project_name'] == project]
        selected_exists = not sel_row.empty
        curr_sel_exists = not self.taiga_projects_df.loc[self.taiga_projects_df['is_selected'] == 1].empty

        if selected_exists:
            if curr_sel_exists:
                self.taiga_projects_df.loc[self.taiga_projects_df['is_selected'] == 1, 'is_selected'] = 0 

            self.taiga_projects_df.loc[self.taiga_projects_df['project_name'] == project, 'is_selected'] = 1
            self.db.df_to_table('taiga_projects', self.taiga_projects_df)
            self._set_linked_taiga_project(sel_row['id'].iloc[0], sel_row['project_name'].iloc[0], sel_row['project_owner'].iloc[0], sel_row['project_slug'].iloc[0])
            return "Success", f"Successfully linked project '{project}'"
        else:
            return "ERROR", f"No project by name '{project}' exists"

    def taiga_data_ready(self):
        return self.taiga_data_available

    def load_taiga_credentials(self):
        return self._get_site_credentials_from_db(TAIGA)[0:2]
    
    def _update_taiga_credentials(self, uname=None, pwd=None):
        is_success = self._update_user_credentials(TAIGA, username=uname, pwd=pwd)
        return is_success
    
    def update_taiga_csv_urls(self, us_url='NULL', task_url='NULL'):
        self.us_report_url = us_url if us_url != 'NULL' else None
        self.task_report_url = task_url if task_url != 'NULL' else None

        return self.db.update('taiga_csv_urls', {'durl': us_url}, {'dname': 'user_story'}) \
            and self.db.update('taiga_csv_urls', {'durl': task_url}, {'dname': 'task'})
    
    def get_taiga_csv_urls(self):
        us_url = task_url = None
        results = self.db.select('taiga_csv_urls')
        for entry in results:
            if entry:
                dname = entry[0]
                if dname == 'user_story':
                    us_url = entry[1]
                elif dname == 'task':
                    task_url = entry[1]
            
        return us_url, task_url
    
    def clear_taiga_data(self):
        tables = ['members', 'sprints', 'userstories', 'tasks']
        for name in tables:
            self.db.clear_table(name)

        self.sprints_df = self.members_df = self.us_df \
            = self.tasks_df = None
        self.taiga_data_available = False

    def clear_taiga_link(self):
        self.clear_taiga_data()
        self.db.clear_table('taiga_projects')
        self._update_taiga_credentials()
        self.update_taiga_csv_urls()
        self.ts.clear_linked_data()

        self.taiga_projects_df = self.sel_pid = self.sel_project_name \
            = self.sel_project_owner = None
        self.project_selected = False
    
    #### Git
    ####===========================================================================

    def commit_data_ready(self) -> bool:
        return self.commits_df is not None and len(self.commits_df) > 0
    
    def pull_all_repos(self):
        for acct in self.git_accts.keys():
            self.pull_repos(acct)
    
    def pull_repos(self, nickname):
        self._update_repos(self.gs.get_repos(nickname))

    def get_avail_repos(self):
        repos = []
        if self.repos is not None:
            repos_df = self.repos.loc[self.repos['is_linked'] == 0]
            for _, row in repos_df.iterrows():
                repos.append([row['site_nickname'], row['repo_name']])
        return repos
    
    def get_linked_repos(self):
        linked = []
        if self.repos is not None:
            linked_df = self.repos.loc[self.repos['is_linked'] == 1]
            for _, row in linked_df.iterrows():
                linked.append([row['site_nickname'], row['repo_name']])
        return linked
    
    def clear_commit_data(self):
        self.commits_df = None
        for repo in self.repos['repo_name'].tolist():
            self._update_latest_commit_date(repo, None)
        self.db.clear_table('commits')

    def get_commits_df(self):
        return self.commits_df

    def import_commit_data(self):
        linked_df = self.repos.loc[self.repos['is_linked'] == 1]
        for _, row in linked_df.iterrows():
            nname = row['site_nickname']
            repo = row['repo_name']
            repo_id = row['id']
            latest_dt = row['last_commit_dt']
            if pd.isna(latest_dt):
                latest_dt = None

            for res, data in self.gs.import_commit_data(nname, repo, repo_id, latest_dt):
                if res == 'In Progress':
                    yield res, f'[{nname}] {data}'
                elif res == 'Complete':
                    df = data[1]
                    if df is not None and len(df) > 0:
                        self.update_commit_df(df)
                        latest_commit_date = df['utc_datetime'].max()
                        latest_commit_str = latest_commit_date.strftime('%Y-%m-%dT%H:%M:%SZ')
                        self._update_latest_commit_date(repo, latest_commit_str)
                    yield res, f'[{nname}] {data[0]}'
    
    ## Data Manipulation
    ##=============================================================================

    def _inv_val_format(self, df: pd.DataFrame):
        df = df.replace(['', 'None', 'nan', 'NaN', np.nan, None], pd.NA)
        
    def update_df(self, curr_df: pd.DataFrame, new_df: pd.DataFrame, col_to_index='id', cols=None):
        try:
            if curr_df is not None and not curr_df.empty:
                curr_df_copy = curr_df.copy(deep=True)
                new_df_copy = new_df.copy(deep=True)

                curr_df_copy = pd.concat([curr_df_copy, new_df_copy]).drop_duplicates([col_to_index], keep='first')

                curr_df_copy = curr_df_copy.set_index(col_to_index)
                new_df_copy = new_df_copy.set_index(col_to_index)

                if cols is not None:
                    curr_df_copy.update(new_df_copy[cols])
                else:
                    curr_df_copy.update(new_df_copy)
                curr_df_copy = curr_df_copy.reset_index()
                return curr_df_copy
            else:
                curr_df = new_df
                return curr_df
        except Exception as e:
            print(e)
            return curr_df
        
    #### Taiga
    ####===========================================================================
    
    def _update_projects_df(self, new_df):
        def format_df(df: pd.DataFrame) -> pd.DataFrame:
            self._inv_val_format(df)
            df['id'] = df['id'].astype(pd.Int64Dtype())
            
            df = df.dropna(how='all')
            df = df.drop_duplicates(subset=['id'], keep='first').reset_index(drop=True)
            df = df.sort_values(by='id', ascending=True)
            return df

        if new_df is not None and len(new_df) > 0:
            self.taiga_projects_df = self.update_df(self.taiga_projects_df, format_df(new_df))
            self.db.df_to_table('taiga_projects', self.taiga_projects_df)
            self.taiga_projects_df = format_df(self.db.table_to_df('taiga_projects'))

    def _update_sprints_df(self, new_df: pd.DataFrame):
        def to_table_format(df: pd.DataFrame) -> pd.DataFrame:
            df['sprint_start'] = pd.to_datetime(df['sprint_start']).dt.strftime('%m/%d/%Y')
            df['sprint_end'] = pd.to_datetime(df['sprint_end']).dt.strftime('%m/%d/%Y')
            return df

        def format_df(df: pd.DataFrame) -> pd.DataFrame:
            self._inv_val_format(df)
            df['id'] = df['id'].astype(pd.Int64Dtype())
            df['sprint_start'] = pd.to_datetime(df['sprint_start'])
            df['sprint_end'] = pd.to_datetime(df['sprint_end'])
            
            df = df.dropna(how='all')
            df = df.drop_duplicates(subset=['id'], keep='first').reset_index(drop=True)
            df = df.sort_values(by='sprint_start', ascending=True)
            return df

        if new_df is not None and len(new_df) > 0:
            self.sprints_df = self.update_df(self.sprints_df, format_df(new_df))
            self.db.df_to_table('sprints', to_table_format(self.sprints_df))
            self.sprints_df = format_df(self.db.table_to_df('sprints'))


    def _update_members_df(self, new_df: pd.DataFrame):
        def format_df(df: pd.DataFrame) -> pd.DataFrame:
            self._inv_val_format(df)
            df['id'] = df['id'].astype(pd.Int64Dtype())
            
            df = df.dropna(how='all')
            df = df.drop_duplicates(subset=['username'], keep='first').reset_index(drop=True)
            df = df.sort_values(by='id', ascending=True)
            return df
        
        if new_df is not None and len(new_df) > 0:
            self.members_df = self.update_df(self.members_df, format_df(new_df), 'username')
            self.db.df_to_table('members', self.members_df)
            self.members_df = format_df(self.db.table_to_df('members'))
    
    def update_us_df(self, new_df: pd.DataFrame, cols=None):
        def format_df(df: pd.DataFrame) -> pd.DataFrame:
            self._inv_val_format(df)
            df['id'] = df['id'].astype(pd.Int64Dtype())
            df['us_num'] = df['us_num'].astype(pd.Int64Dtype())
            df['points'] = df['points'].astype(pd.Int64Dtype())
            df['is_complete'] = df['is_complete'].astype(pd.BooleanDtype())
            df['points'] = df['points'].replace(pd.NA, 0)
            
            df = df.dropna(how='all')
            df = df.drop_duplicates(subset=['id'], keep='first').reset_index(drop=True)
            df = df.sort_values(by=['sprint', 'us_num'], ascending=True, na_position='last')
            return df
        
        if new_df is not None and len(new_df) > 0:
            self.us_df = self.update_df(self.us_df, format_df(new_df))
            self.db.df_to_table('userstories', self.us_df)
            self.us_df = format_df(self.db.table_to_df('userstories'))

        self.taiga_data_available = self.us_df is not None and len(self.us_df) > 0 and self.tasks_df is not None and len(self.tasks_df) > 0

    def update_tasks_df(self, new_df: pd.DataFrame, cols=['task_num', 'is_complete', 'us_num', 'assignee', 'task_subject']):
        def format_df(df: pd.DataFrame) -> pd.DataFrame:
            df['id'] = df['id'].astype(pd.Int64Dtype())
            df['task_num'] = df['task_num'].astype(pd.Int64Dtype())
            df['us_num'] = df['us_num'].astype(pd.Int64Dtype())
            df['is_coding'] = df['is_coding'].astype(pd.BooleanDtype())
            df['is_complete'] = df['is_complete'].astype(pd.BooleanDtype())

            self._inv_val_format(df)
            df = df.dropna(how='all')
            df = df.drop_duplicates(subset=['id'], keep='first').reset_index(drop=True)
            df = df.sort_values(by=['sprint', 'us_num', 'task_num'], ascending=True)
            return df
        
        if new_df is not None and len(new_df) > 0:
            self.tasks_df = self.update_df(self.tasks_df, format_df(new_df), cols=cols)
            self.db.df_to_table('tasks', self.tasks_df)
            self.tasks_df = format_df(self.db.table_to_df('tasks'))

        self.taiga_data_available = self.us_df is not None and len(self.us_df) > 0 and self.tasks_df is not None and len(self.tasks_df) > 0

    ## Data Importing
    ##=============================================================================
    #### Taiga
    ####===========================================================================

    def _pull_taiga_projects(self):
        projects = self.ts.get_watched_projects()
        if projects is not None and not projects.empty:
             self._update_projects_df(projects)

    def wait_for_projects(self):
        self._pull_taiga_projects()
        return True
    
    def _process_taiga_data(self, sprint_df, member_df, us_df, tasks_df):
        self._update_sprints_df(sprint_df)
        self._update_members_df(member_df)
        self.update_us_df(us_df)
        self.update_tasks_df(tasks_df)

    def taiga_import_by_api(self):
        if self.project_selected and self.ts.token_set():
            try:
                sprints_df, members_df, us_df, task_df = self.ts.import_data_by_api(self.sel_pid)
                self._process_taiga_data(sprints_df, members_df, us_df, task_df)
                return 'Success', f'Successfully imported Taiga data by API'
            except Exception as e:
                return 'Error', f'Failed to import Taiga data by API - {e}'

    def taiga_import_by_urls(self, us_url, tasks_url):
        if us_url and tasks_url:
            try:
                sprints_df, members_df, us_df, task_df = self.ts._import_data_by_urls(us_url, tasks_url)
                self._process_taiga_data(sprints_df, members_df, us_df, task_df)
                return 'Success', f'Successfully imported Taiga data by URLs'
            except Exception as e:
                return 'Error', f'Failed to import Taiga data by URLs - {e}'

    def taiga_import_by_files(self, us_fp, tasks_fp):
        if us_fp and tasks_fp:
            try:   
                sprints_df, members_df, us_df, task_df = self.ts._import_data_by_files(us_fp, tasks_fp)
                self._process_taiga_data(sprints_df, members_df, us_df, task_df)
                return 'Success', f'Successfully imported Taiga data by File'
            except Exception as e:
                return 'Error', f'Failed to import Taiga data by File - {e}'
            
    #### Git
    ####===========================================================================

    def _update_latest_commit_date(self, repo, date):
        self.repos.loc[self.repos['repo_name'] == repo, 'last_commit_dt'] = date
        self._update_repos(self.repos, cols=['last_commit_dt'])

    def update_commit_df(self, new_df : pd.DataFrame):
        def to_table_format(df) -> pd.DataFrame:
            # df['utc_datetime'] = pd.to_datetime(df['utc_datetime']).dt.strftime('%Y-%m-%dT%H:%M:%SZ')
            return df

        def format_df(df: pd.DataFrame) -> pd.DataFrame:
            self._inv_val_format(df)
            df['id'] = df['id'].astype(pd.StringDtype())
            df['task_num'] = df['task_num'].astype(pd.Int64Dtype())
            df['utc_datetime'] = pd.to_datetime(df['utc_datetime'])

            df = df.dropna(how='all')
            df = df.drop_duplicates(subset=['id', 'repo_name'], keep='first').reset_index(drop=True)
            df = df.sort_values(by='utc_datetime', ascending=True)
            return df

        if new_df is not None and len(new_df) > 0:
            self.commits_df = self.update_df(self.commits_df, format_df(new_df))
            self.db.df_to_table('commits', to_table_format(self.commits_df))
            self.commits_df = format_df(self.db.table_to_df('commits'))
        self.commit_data_available = self.commits_df is not None and len(self.commits_df) > 0

    def _update_repos(self, new_df: pd.DataFrame, cols=['repo_name', 'owner_name']):
        def format_df(df: pd.DataFrame) -> pd.DataFrame:
            self._inv_val_format(df)
            df['id'] = df['id'].astype(pd.Int64Dtype())

            df = df.dropna(how='all')
            df = df.drop_duplicates(subset=['id', 'site_nickname'], keep='first').reset_index(drop=True)
            df = df.sort_values(by=['site_nickname'], ascending=True)
            return df

        if new_df is not None and len(new_df) > 0:
            self.repos = self.update_df(self.repos, format_df(new_df), cols=cols)
            self.db.df_to_table('repos', self.repos)
            self.repos = format_df(self.db.table_to_df('repos'))
        
    ## Data File reading/writing
    ##=============================================================================

    def _create_new_wb(self, filename, sheets=None):
        if os.path.exists(filename):
            os.remove(filename)
        
        wb = opyxl.Workbook()
        wb.create_sheet("Master")

        if sheets is not None:
            for sheet in sheets:
                wb.create_sheet(sheets)

        for sheet in wb.sheetnames:
            if sheets is not None:
                if sheet not in sheets and sheet != 'Master':
                    del wb[sheet]
            else:
                if sheet != 'Master':
                    del wb[sheet]
            
        wb.save(filename)
        wb.close()

    def _parsed_data_to_spreadsheet(self, df, writer, sheet):
        df.to_excel(writer, sheet_name=sheet, index=False)

    def write_to_csv(self, filepath: Type[str], df: Type[pd.DataFrame]):
        subdirectories = filepath.replace('.', '').split('/')
        curr_dir_level = '.'
        for i in range(len(subdirectories) - 1):
            dir = subdirectories[i]
            curr_dir_level += f'/{dir}'

            if not os.path.exists(curr_dir_level):
                os.makedirs(curr_dir_level)

        df.to_csv(filepath, index=False)

    def write_to_excel(self, filepath, df, header_filter=None, sheets=None, sheet_headers=None):
        self._create_new_wb(filepath, sheets)
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            self._parsed_data_to_spreadsheet(df, writer, 'Master')

            if sheets is not None:
                for sheet in sheets:
                    sheet_df = df[df[header_filter == sheet]][sheet_headers]
                    self._parsed_data_to_spreadsheet(sheet_df, writer, sheet)
    
    def remove_file(self, filepath):
        if os.path.exists(filepath):
            os.remove(filepath)
    
    ## Data retrieval/setting
    ##=============================================================================

    def format_spreadsheet(self, filepath):
        if not os.path.exists(filepath):
            return ['Error', 'File does not exist.']
    
        try:
            wb = opyxl.load_workbook(filepath)
            for sheet in wb.worksheets:
                col_dims = dict()
                colidx_range = [i for i in range(sheet.min_column, sheet.max_column + 1)]
                for colidx in colidx_range:
                    column = sheet.cell(1, colidx).column_letter
                    link_col = False
                    date_col = False
                    int_col = False

                    header = sheet.cell(1, colidx).value
                    if 'link' in header.lower():
                        link_col = True
                    elif 'date' in header.lower():
                        date_col = True
                    elif header in ['Task #', 'User Story', 'Points']:
                        int_col = True

                    col_dims[column] = len(str(header))
                    for i in range(2, sheet.max_row + 1):
                        cell = sheet.cell(i, colidx)

                        try:
                            if link_col or isinstance(cell.value, str) and cell.value.startswith('=HYPERLINK('):
                                # Extract URL from the formula
                                url_start = cell.value.find('"') + 1
                                url_end = cell.value.find('"', url_start)
                                url = cell.value[url_start:url_end] if url_start > 0 and url_end > url_start else ""

                                if url is not None and url != '':
                                    # Extract Friendly Text (if available)
                                    text_start = cell.value.find('"', url_end + 1) + 1
                                    text_end = cell.value.find('"', text_start)
                                    friendly_text = cell.value[text_start:text_end] if text_start > 0 and text_end > text_start else url

                                    # Set the hyperlink in the cell
                                    cell.value = friendly_text  # Display text
                                    cell.hyperlink = url  # Set hyperlink
                                    cell.font = styles.Font(underline='single', color='0000FF') # Apply Excel hyperlink style 
                                    cell_text = friendly_text
                            elif date_col:
                                date = pd.to_datetime(cell.value)
                                cell.value = date
                                cell.number_format = 'm/d/yyyy'
                                cell_text = str(cell.value)
                            elif int_col:
                                cell.value = int(cell.value)
                                cell.number_format = styles.numbers.FORMAT_NUMBER
                                cell_text = str(cell.value)
                            else:
                                cell_text = str(cell.value)
                        except:
                            cell_text = ''

                        col_dims[column] = max(col_dims.get(column), len(str(cell_text)))
                for col, val in col_dims.items():
                    sheet.column_dimensions[col].width = val + 5

            wb.save(filepath)
            return ['Success']

        except Exception as e:
            print(f'Failed to format spreadsheet - {e}')
            return ['Error', str(e)]
    
    ## Data Formatting for Reports
    ##=============================================================================
    
    def _generate_hyperlink(self, url, text):
        return f'=HYPERLINK("{url}", "{text}")' if pd.notna(url) else pd.NA
    
    def _generate_task_excel_entry(self, task_num, url=None, preceding_text=None):
        if pd.notna(task_num):
            taiga_tasks = []

            if self.taiga_data_ready():
                taiga_tasks = self.tasks_df['task_num'].tolist()

            try:
                task_num = int(task_num)
            except:
                pass

            if task_num in taiga_tasks:
                if preceding_text is not None:
                    text = f'{preceding_text} Task-{task_num}'
                else:
                    text = f'Task-{task_num}'

                if url is None or url == '':
                    url = f'{self._get_taiga_base_url()}{task_num}'
                text = self._generate_hyperlink(url, text)      
            else:
                text = f'Task-{task_num} (No Link)'

            return text
        return pd.NA
    
    def _generate_us_entry(self, us_num):
        if us_num == 'Storyless' or pd.isna(us_num):
            return 'Storyless'
        else:
            return f'US-{us_num}'
        
    
    
    def format_wsr_non_excel(self, df: pd.DataFrame):
        members = df['assignee'].dropna().drop_duplicates().tolist()
        num_mems = len(members)

        data_columns = ['Sprint', 'User Story', 'Points', 'Task', 'Coding?']
        data_columns.extend(members)

        data = []
        for _, row in df.iterrows():
            us_num = row['us_num']
            task_num = row['task_num']
            assigned = row['assignee']

            sprint = row['sprint']
            user_story = int(us_num) if not pd.isna(us_num) else pd.NA
            points = int(row['points'])
            task = int(task_num) if pd.notna(task_num) else pd.NA
            coding = 'TRUE' if row['is_coding'] == 1 else 'FALSE'

            mem_data = [None] * num_mems
            if pd.notna(assigned):
                for idx, mem in enumerate(members):
                    mem_data[idx] = "100%" if assigned == mem else None

            row_data = [sprint, user_story, points, task, coding] + mem_data
            data.append(row_data)

        result_df = pd.DataFrame(data=data, columns=data_columns)
        return result_df
    
    def format_wsr_excel(self, df: pd.DataFrame):
        df['Task'] = df['Task'].apply(lambda x: self._generate_task_excel_entry(x))
        df['User Story'] = df['User Story'].apply(lambda x: self._generate_us_entry(x))
        return df
    
    def format_icr_df_non_excel(self, commit_df: pd.DataFrame, taiga_df: pd.DataFrame = None) -> pd.DataFrame:
        commit_df = commit_df.sort_values(by='utc_datetime', ascending=True)

        data_columns = ['Committer', 'Link to Task', 'Task #', 'Task Status', 'Coding Task?', 'Link to Commit', 'Commit Date', 'Percent Contributed']
        data = [None] * len(commit_df)
        for index, row in commit_df.iterrows():
            
            committer = row['committer']
            task_num = row['task_num']
            if not pd.isna(task_num):
                task_url = f'{self._get_taiga_base_url()}{int(task_num)}'
                task = task_num
                is_complete = taiga_df.loc[taiga_df['task_num'] == task_num, 'is_complete'].iloc[0] if taiga_df is not None else None
                task_status = 'Complete' if is_complete else 'In-Process' if taiga_df is not None else None
                coding_val = taiga_df.loc[taiga_df['task_num'] == task_num, 'is_coding'].iloc[0]
                coding = 'TRUE' if coding_val else 'FALSE' if coding_val is not None else pd.NA
            else:
                task_url = pd.NA
                task = pd.NA
                coding = pd.NA
                task_status = pd.NA

            commit_url = row['commit_url']
            commit_date = row['az_date']

            row_data = [committer, task_url, task, task_status, coding, commit_url, commit_date, '100']
            data[index] = row_data
            
        result_df = pd.DataFrame(data, columns=data_columns)
        result_df['Task #'] = result_df['Task #'].astype(pd.Int64Dtype())
        result_df = result_df.astype(pd.StringDtype())
        result_df = result_df.replace(pd.NA, '')
        return result_df
    
    def format_icr_excel(self, icr_df: pd.DataFrame):
        def format_links(df: pd.DataFrame) -> pd.DataFrame:
            task_commit_dict = dict()

            for idx, row in df.iterrows():
                commit_link_text = 'Link to Commit (No Linked Task)'

                task_num = row['Task #']
                task_url = row['Link to Task']
                commit_url = row['Link to Commit']

                if pd.notna(task_num):
                    commit_count = task_commit_dict.get(task_num)
                    if not commit_count:
                        task_commit_dict[task_num] = 1
                        commit_count = 1

                    commit_link_text = f'Link to Task-{task_num} Commit #{commit_count}'
                    task_commit_dict[task_num] += 1

                task_hyperlink = self._generate_task_excel_entry(task_num=task_num, url=task_url, preceding_text='Link to')
                commit_hyperlink = self._generate_hyperlink(commit_url, commit_link_text)

                df.at[idx, 'Link to Task'] = task_hyperlink
                df.at[idx, 'Link to Commit'] = commit_hyperlink

        format_links(icr_df)
        return icr_df