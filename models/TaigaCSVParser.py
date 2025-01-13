import os
import io
from typing import Type
import pandas as pd
import numpy as np
import openpyxl as opyxl
import requests


class TaigaParsingController:
    us_report_url = None
    task_report_url = None

    us_fp = None
    task_fp = None

    member_list = None
    sprint_list = None
    sprint_dates = None
    us_list = None
    raw_us_df = None
    raw_task_df = None
    formatted_master_df = None

    data_ready = False
    
    def __init__(self, us_url, task_url):
        self.set_us_report_url(us_url)
        self.set_task_report_url(task_url)

    def conv_nan_to_none(self, val):
        if pd.isna(val):
            return None
        else:
            return val

    def set_master_df(self, df):
        self.formatted_master_df = df
        self.data_ready = True

    def load_raw_data(self, master_df, us_df, task_df):
        if master_df is not None:
            master_df.replace(['', 'None', 'nan', 'NaN'], [None, None, None, None], inplace=True)
            self.set_master_df(master_df)
        if us_df is not None:
            self.__format_us_data(us_df)
        if task_df is not None:
            self.__format_task_data(task_df)

    def data_is_ready(self) -> bool:
        return self.data_ready

    ## API/File declaration
    ##=============================================================================

    def set_us_report_url(self, url):
        if url != "" and url is not None:
            self.us_report_url = url

    def get_us_report_url(self):
        return self.us_report_url

    def set_task_report_url(self, url):
        if url != "" and url is not None:
            self.task_report_url = url

    def get_task_report_url(self):
        return self.task_report_url

    def __file_is_valid(self, filename):
        if filename != "" and filename is not None:
            return True
        else:
            return False

    def set_us_fp(self, fp):
        if self.__file_is_valid(fp):
            self.us_fp = fp
            return True
        return False
    
    def get_us_fp(self):
        return self.us_fp

    def set_task_fp(self, fp):
        if self.__file_is_valid(fp):
            self.task_fp = fp
            return True
        return False
    
    def get_task_fp(self):
        return self.task_fp
    
    ## Retrieval of data
    ##=============================================================================

    def __us_data_from_api(self):
        url = self.us_report_url

        if url != "" and url is not None:
            res = requests.get(url)._content
            raw_data = pd.read_csv(io.StringIO(res.decode('utf-8')))
            return self.__format_us_data(raw_data)
        return False

    def __task_data_from_api(self):
        url = self.task_report_url

        if url != "" and url is not None:
            res = requests.get(url)._content
            raw_data = pd.read_csv(io.StringIO(res.decode('utf-8')))
            return self.__format_task_data(raw_data)
        return False

    def retrieve_data_by_api(self):
        task_parse_success = self.__task_data_from_api()
        us_parse_success = self.__us_data_from_api()

        if task_parse_success and us_parse_success:
            self.__format_and_centralize_data()
            self.data_ready = True
        else:
            print('Error retrieving and parsing data through api')

    def __us_data_from_file(self):
        us_fp = self.us_fp
        us_url_set = self.__file_is_valid(us_fp)

        if not us_url_set:
            us_url_set = self.set_us_fp()

        if us_url_set:
            raw_data = None

            if us_fp.split(".")[-1] == "csv":
                raw_data = pd.read_csv(us_fp)
            elif us_fp.split(".")[-1] == "xlsx":
                raw_data = pd.read_excel(us_fp)
            return self.__format_us_data(raw_data)
        return False
    
    def get_raw_us_data(self):
        return self.raw_us_df
    
    def __inv_val_to_none(self, df: Type[pd.DataFrame]):
        df.replace(['', 'None', 'nan', 'NaN', np.nan], [None, None, None, None, None], inplace=True)
    
    def __parse_user_stories(self, df):
        df_to_use = df.copy(deep=True)
        self.__inv_val_to_none(df_to_use)
        df_to_use.replace(False, None, inplace=True)
        df_to_use.dropna(inplace=True)
        df_to_use = df_to_use.drop_duplicates(keep='first').reset_index(drop=True)
        self.us_list = df_to_use['ref'].tolist()
    
    def get_user_stories(self):
        return self.us_list

    def __task_data_from_file(self):
        task_fp = self.task_fp
        task_url_set = self.__file_is_valid(task_fp)

        if not task_url_set:
            task_url_set = self.set_task_fp()

        if task_url_set:
            raw_data = None

            if task_fp.split(".")[-1] == "csv":
                raw_data = pd.read_csv(task_fp)
            elif task_fp.split(".")[-1] == "xlsx":
                raw_data = pd.read_excel(task_fp)
            return self.__format_task_data(raw_data)
        return False
    
    def get_raw_task_data(self):
        return self.raw_task_df

    def retrieve_data_by_file(self):
        task_parse_success = self.__task_data_from_file()
        us_parse_success = self.__us_data_from_file()

        if task_parse_success and us_parse_success:
            self.__format_and_centralize_data()
            self.data_ready = True
        else:
            print('Error retrieving and parsing data from files')

    def clear_data(self):
        self.member_list = None
        self.sprints_df = None
        self.raw_us_df = None
        self.raw_task_df = None
        self.formatted_master_df = None

    def __parse_members(self, df):
        df_to_use = df.copy(deep=True)
        self.__inv_val_to_none(df_to_use)
        df_to_use.dropna(inplace=True)
        df_to_use = df_to_use.drop_duplicates(keep='first').reset_index(drop=True)
        self.member_list = df_to_use.tolist()

    def get_members(self):
        return self.member_list

    def __parse_sprints(self, df):
        df_to_use = df.copy(deep=True)
        self.__inv_val_to_none(df_to_use)
        df_to_use.dropna(subset=['sprint', 'sprint_estimated_start', 'sprint_estimated_finish'], inplace=True)
        self.sprints_df = df_to_use.drop_duplicates(subset=['sprint'], keep='first').reset_index(drop=True)

    def get_sprints(self):
        return self.sprints_df['sprint'].tolist()
    
    ## Data preparation
    ##=============================================================================
    
    def __format_us_data(self, raw_df):
        if raw_df is not None:
                self.raw_us_df = raw_df[['id', 'ref', 'is_closed', 'sprint', 'sprint_estimated_start', 'sprint_estimated_finish', 'total-points']]
                self.__parse_sprints(raw_df[['sprint', 'sprint_estimated_start', 'sprint_estimated_finish']])
                self.__parse_user_stories(raw_df[['ref', 'is_closed']])
                return True
        return False
    
    def __format_task_data(self, raw_df):
        if raw_df is not None:
            raw_df = raw_df[['id', 'ref', 'is_closed', 'subject', 'user_story', 'sprint', 'assigned_to']]
            raw_df.sort_values(['sprint'], ascending=[True], inplace=True)
            self.raw_task_df = raw_df
            self.__parse_members(raw_df['assigned_to'])
            return True
        return False

    def __make_hyperlink(self, task_num):
        base_url = 'https://tree.taiga.io/project/dlsmallw-group-8-asu-capstone-natural-language-processing-for-decolonizing-harm-reduction-literature/task/{}'
        return '=HYPERLINK("%s", "Task-%s")' % (base_url.format(task_num), task_num)
    
    def __get_sprint_date(self, sprint):
        start = self.sprints_df[self.sprints_df['sprint'] == sprint]['sprint_estimated_start']
        end = self.sprints_df[self.sprints_df['sprint'] == sprint]['sprint_estimated_finish']
        
        start_date = self.__extract_date(start).date()
        end_date = self.__extract_date(end).date()

        return start_date.strftime('%m/%d/%Y'), end_date.strftime('%m/%d/%Y')

    def __extract_date(self, date_obj) -> pd.Timestamp | None:
        idx_obj = date_obj.values
        raw_val = idx_obj[0] if len(idx_obj) > 0 else None
        return pd.to_datetime(raw_val, format='%Y-%m-%d') if raw_val is not None else None

    def __format_and_centralize_data(self):
        data_columns = ['sprint', 'sprint_start', 'sprint_end', 'user_story', 'points', 'task', 'assigned_to', 'coding', 'subject']

        us_df = self.raw_us_df
        task_df = self.raw_task_df

        all_data = [0] * len(task_df)

        for index, row in task_df.iterrows():
            us_num = row['user_story']
            us_row = us_df.loc[us_df['ref'] == us_num]

            
            sprint = row['sprint']
            sprint_start, sprint_end = self.__get_sprint_date(sprint)

            user_story = int(us_num) if pd.notnull(us_num) else None
            points = int(us_row['total-points'].values[0] if pd.notnull(us_num) else 0)
            task = int(row['ref'])
            assigned = row['assigned_to'] if pd.notnull(row['assigned_to']) else 'Unassigned'
            coding = ""
            subject = row['subject']
            
            data_row = [sprint, sprint_start, sprint_end, user_story, points, task, assigned, coding, subject]
            all_data[index] = data_row

        self.set_master_df(pd.DataFrame(all_data, columns=data_columns))

    def __format_df_for_excel(self, df):
        members = self.get_members()
        num_mems = len(members)

        data_columns = ['sprint', 'user_story', 'points', 'task', 'coding']
        data_columns.extend(members)

        data = [None] * len(df)
        for index, row in df.iterrows():
            us_num = row['user_story']
            task_num = row['task']
            assigned = row['assigned_to']

            sprint = row['sprint']
            user_story = f'US-{int(us_num)}' if us_num is not None else 'Storyless'
            points = int(row['points'])
            task = self.__make_hyperlink(task_num)
            coding = row['coding']

            mem_data = [None] * num_mems
            i = 0
            for mem in members:
                mem_data[i] = "100%" if assigned == mem else None
                i += 1

            row_data = [sprint, user_story, points, task, coding]
            row_data.extend(mem_data)
            data[index] = row_data
            
        result_df = pd.DataFrame(data, columns=data_columns)
        return result_df
    

    
    ## File Writing
    ##=============================================================================

    def __create_new_wb(self, filename):
        if (os.path.exists(filename)):
            os.remove(filename)
        
        wb = opyxl.Workbook()
        wb.create_sheet("All_Data")

        sprints = self.get_sprints()

        for sprint in sprints:
            wb.create_sheet(sprint)

        for sheet in wb.sheetnames:
            if sheet not in sprints and sheet != "All_Data":
                del wb[sheet]

        wb.save(filename)

    def __parsed_data_to_spreadsheet(self, df, writer, sheet):
        df.to_excel(writer, sheet_name=sheet)

    def write_data(self, filename):
        sprints = self.get_sprints()
        self.__create_new_wb(filename, sprints)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            master_df = self.get_master_df()
            self.__parsed_data_to_spreadsheet(self.__format_df_for_excel(master_df), writer, "All_Data")
            
            for sprint in sprints:
                sprint_df = master_df.loc[master_df['sprint'] == sprint]
                self.__parsed_data_to_spreadsheet(
                    self.__format_df_for_excel(sprint_df.sort_values(['sprint', 'user_story', 'task'], ascending=[True, True, True])), 
                    writer, sprint)
                
    def get_master_df(self):
        return self.formatted_master_df.sort_values(['sprint', 'user_story', 'task'], ascending=[True, True, True])